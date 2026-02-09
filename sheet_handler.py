"""
Sheet Handler Module
Handles reading and managing multiple sheets in Excel files
"""

import pandas as pd
from openpyxl import load_workbook


def get_sheet_names(file_path):
    """
    Get all sheet names from an Excel file
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        list: List of sheet names
    """
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        print(f"  Warning: Could not read sheet names from {file_path.name}: {str(e)}")
        return []


def read_all_sheets(file_path):
    """
    Read all sheets from an Excel file
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        dict: Dictionary with sheet_name as key and DataFrame as value
    """
    try:
        # Read all sheets
        all_sheets = pd.read_excel(file_path, sheet_name=None)
        return all_sheets
    except Exception as e:
        print(f"  Error reading sheets from {file_path.name}: {str(e)}")
        return {}


def read_specific_sheets(file_path, sheet_names):
    """
    Read specific sheets from an Excel file
    
    Args:
        file_path: Path to Excel file
        sheet_names: List of sheet names to read
    
    Returns:
        dict: Dictionary with sheet_name as key and DataFrame as value
    """
    sheets = {}
    
    for sheet_name in sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            sheets[sheet_name] = df
        except Exception as e:
            print(f"  Warning: Could not read sheet '{sheet_name}' from {file_path.name}: {str(e)}")
    
    return sheets


def read_excel_sheets_with_formulas(file_path, sheet_names=None):
    """
    Read Excel file sheets and extract formulas
    
    Args:
        file_path: Path to Excel file
        sheet_names: List of specific sheet names, or None for all sheets
    
    Returns:
        dict: {sheet_name: {'dataframe': df, 'formulas': formulas_dict}}
    """
    result = {}
    
    try:
        wb = load_workbook(file_path, data_only=False)
        
        # Determine which sheets to process
        if sheet_names is None:
            sheets_to_process = wb.sheetnames
        else:
            sheets_to_process = [s for s in sheet_names if s in wb.sheetnames]
        
        for sheet_name in sheets_to_process:
            # Read data as DataFrame
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Extract formulas
            formulas = {}
            ws = wb[sheet_name]
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2)):  # Skip header
                for col_idx, cell in enumerate(row):
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas[(row_idx, col_idx)] = cell.value
            
            result[sheet_name] = {
                'dataframe': df,
                'formulas': formulas
            }
        
        wb.close()
        
    except Exception as e:
        print(f"  Error reading sheets with formulas from {file_path.name}: {str(e)}")
    
    return result


def compare_sheet_structures(sheets_A, sheets_B, filename):
    """
    Compare sheet structures between two files
    
    Args:
        sheets_A: Dict of sheets from Folder A
        sheets_B: Dict of sheets from Folder B
        filename: Name of the file
    
    Returns:
        dict: Structure comparison info
    """
    comparison = {
        'filename': filename,
        'sheets_match': set(sheets_A.keys()) == set(sheets_B.keys()),
        'sheets_only_in_A': list(set(sheets_A.keys()) - set(sheets_B.keys())),
        'sheets_only_in_B': list(set(sheets_B.keys()) - set(sheets_A.keys())),
        'common_sheets': list(set(sheets_A.keys()) & set(sheets_B.keys())),
        'total_sheets_A': len(sheets_A),
        'total_sheets_B': len(sheets_B)
    }
    
    return comparison


def get_sheet_selection(all_available_sheets, compare_all_sheets=True, specific_sheets=None):
    """
    Determine which sheets to compare based on configuration
    
    Args:
        all_available_sheets: List of all sheet names in file
        compare_all_sheets: Whether to compare all sheets
        specific_sheets: List of specific sheet names (if compare_all_sheets=False)
    
    Returns:
        list: Sheet names to compare
    """
    if compare_all_sheets:
        return all_available_sheets
    
    if specific_sheets:
        # Return only sheets that exist in the file
        return [s for s in specific_sheets if s in all_available_sheets]
    
    # Default: return first sheet only
    return [all_available_sheets[0]] if all_available_sheets else []


def format_sheet_info(sheets_dict):
    """
    Format sheet information for logging
    
    Args:
        sheets_dict: Dictionary of sheets
    
    Returns:
        str: Formatted string with sheet info
    """
    if not sheets_dict:
        return "No sheets found"
    
    info_lines = []
    for sheet_name, sheet_data in sheets_dict.items():
        if isinstance(sheet_data, dict) and 'dataframe' in sheet_data:
            df = sheet_data['dataframe']
            info_lines.append(f"    • '{sheet_name}': {df.shape[0]} rows × {df.shape[1]} cols")
        elif isinstance(sheet_data, pd.DataFrame):
            info_lines.append(f"    • '{sheet_name}': {sheet_data.shape[0]} rows × {sheet_data.shape[1]} cols")
    
    return "\n".join(info_lines)
