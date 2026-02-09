"""
File Reading & Preprocessing Module
Handles reading Excel files and preprocessing (whitespace stripping, formula extraction)
"""

import pandas as pd
from openpyxl import load_workbook
from column_processor import remove_blank_columns


def read_excel_with_formulas(file_path):
    """Read Excel file and extract formulas"""
    df = pd.read_excel(file_path)
    formulas = {}
    
    try:
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2)):  # Skip header
            for col_idx, cell in enumerate(row):
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas[(row_idx, col_idx)] = cell.value
        
        wb.close()
    except Exception as e:
        print(f"  Warning: Could not extract formulas: {str(e)}")
    
    return df, formulas


def strip_dataframe_whitespace(df):
    """Strip leading/trailing whitespace from all string cells and column names"""
    # Strip column names
    df.columns = df.columns.str.strip() if hasattr(df.columns, 'str') else [str(c).strip() for c in df.columns]
    
    # Strip all string values in cells
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
    
    return df


def read_folder_files(folder_path, strip_whitespace=True, compare_formulas=False, 
                     delete_blank_columns=False, compare_all_sheets=True, specific_sheets=None):
    """
    Read all Excel files in a folder (with multi-sheet support)
    
    Args:
        folder_path: Path object to folder
        strip_whitespace: Whether to strip whitespace
        compare_formulas: Whether to extract formulas
        delete_blank_columns: Whether to remove blank columns
        compare_all_sheets: Whether to read all sheets or specific ones
        specific_sheets: List of specific sheet names (if compare_all_sheets=False)
    
    Returns:
        dict: Dictionary with filename as key and file data as value
    """
    from sheet_handler import read_all_sheets, read_specific_sheets, read_excel_sheets_with_formulas, get_sheet_selection
    
    files_data = {}
    
    excel_files = [f for f in folder_path.iterdir() 
                   if f.is_file() and f.suffix.lower() in ['.xlsx', '.xls']]
    
    for file_path in sorted(excel_files):
        try:
            # Read sheets based on configuration
            if compare_formulas:
                # Read with formulas
                sheets_to_read = None if compare_all_sheets else specific_sheets
                sheets_dict = read_excel_sheets_with_formulas(file_path, sheets_to_read)
            else:
                # Read without formulas
                if compare_all_sheets:
                    sheets_dict = read_all_sheets(file_path)
                    # Convert to consistent format
                    sheets_dict = {name: {'dataframe': df, 'formulas': None} 
                                  for name, df in sheets_dict.items()}
                else:
                    sheets_to_read = specific_sheets if specific_sheets else [pd.ExcelFile(file_path).sheet_names[0]]
                    raw_sheets = read_specific_sheets(file_path, sheets_to_read)
                    sheets_dict = {name: {'dataframe': df, 'formulas': None} 
                                  for name, df in raw_sheets.items()}
            
            # Process each sheet
            processed_sheets = {}
            all_removed_columns = {}
            
            for sheet_name, sheet_data in sheets_dict.items():
                df = sheet_data['dataframe']
                formulas = sheet_data.get('formulas')
                
                # Strip whitespace if enabled
                if strip_whitespace:
                    df = strip_dataframe_whitespace(df)
                
                # Remove blank columns if enabled
                removed_cols = []
                if delete_blank_columns:
                    df, removed_cols = remove_blank_columns(df, f"{file_path.name} - {sheet_name}")
                    if removed_cols:
                        all_removed_columns[sheet_name] = removed_cols
                
                processed_sheets[sheet_name] = {
                    'dataframe': df,
                    'formulas': formulas,
                    'columns': df.columns.tolist(),
                    'shape': df.shape,
                    'removed_blank_columns': removed_cols
                }
            
            files_data[file_path.name] = {
                'sheets': processed_sheets,
                'sheet_names': list(processed_sheets.keys()),
                'total_sheets': len(processed_sheets)
            }
            
            print(f"✓ {file_path.name}")
            print(f"  - Sheets ({len(processed_sheets)}): {list(processed_sheets.keys())}")
            from sheet_handler import format_sheet_info
            print(format_sheet_info(processed_sheets))
            
        except Exception as e:
            print(f"✗ Error reading {file_path.name}: {str(e)}")
    
    return files_data
